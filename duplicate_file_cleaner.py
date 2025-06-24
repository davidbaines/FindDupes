import argparse
import os
from pathlib import Path
from tqdm import tqdm
import send2trash
import shutil  # New import for copying
from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl
import json

from duplicate_file_finder import DEFAULT_REPORT_FILE, files_are_identical


def _read_tasks_from_json(report_path: Path, root_dir: Path, copy_destination: Path | None) -> tuple[list, dict]:
    """Reads tasks and summary from a JSON report file."""
    tasks = []
    with open(report_path, 'r') as f:
        data = json.load(f)

    summary = data.get("summary", {})
    
    for item in tqdm(data.get("files_to_process", []), desc="Reading JSON report"):
        if item.get("type") == "duplicate_group":
            keep_index = item["keep_index"]
            file_paths_str = item["files"]
            file_to_keep_rel = file_paths_str[keep_index]
            file_to_keep_abs = root_dir / file_to_keep_rel

            if copy_destination:
                if file_to_keep_abs.exists():
                    tasks.append(("copy", file_to_keep_abs, copy_destination / file_to_keep_rel, 0))
            else: # Deletion mode
                for idx, path_str in enumerate(file_paths_str):
                    if idx == keep_index:
                        continue
                    file_to_process = root_dir / path_str
                    if file_to_process.exists():
                        tasks.append(("delete", file_to_process, None, 0))

        elif item.get("type") == "unique_file":
            if copy_destination: # Unique files are only relevant for copy mode
                file_to_keep_rel = item["file"]
                file_to_keep_abs = root_dir / file_to_keep_rel
                if file_to_keep_abs.exists():
                    tasks.append(("copy", file_to_keep_abs, copy_destination / file_to_keep_rel, 0))
    return tasks, summary


def _read_tasks_from_xlsx(report_path: Path, root_dir: Path, copy_destination: Path | None) -> tuple[list, dict]:
    """Reads tasks and summary from an XLSX report file."""
    tasks = []
    summary = {}
    workbook = openpyxl.load_workbook(report_path)

    # Try to read summary info
    try:
        summary_sheet = workbook["Summary"]
        summary = {
            "total_files_to_keep": summary_sheet["B2"].value,
            "total_size_to_keep_gb": summary_sheet["B4"].value
        }
    except KeyError:
        pass # No summary sheet, that's fine.

    # Find the correct sheet, preferring the new name but falling back to the old one.
    sheet_name = "Files to Process"
    if sheet_name not in workbook.sheetnames:
        if "Duplicates" in workbook.sheetnames:
            sheet_name = "Duplicates"
            print(f"Info: Sheet 'Files to Process' not found. Using existing sheet 'Duplicates'.")
        else:
            print(f"Error: Could not find a 'Files to Process' or 'Duplicates' sheet in the report.")
            return [], {}
    sheet = workbook[sheet_name]

    total_rows = sheet.max_row - 1 if sheet.max_row > 1 else 0
    row_iterator = sheet.iter_rows(min_row=2, values_only=True)
    for i, row_data in enumerate(tqdm(row_iterator, total=total_rows, desc="Reading XLSX report"), start=2):
        try:
            keep_index_str = str(row_data[0]).strip()
            if not keep_index_str or not keep_index_str.isdigit():
                tqdm.write(f"Skipping row {i}: 'Keep' column is not a valid number ('{row_data[0]}').")
                continue

            keep_index = int(keep_index_str) - 1  # Convert to 0-based index
            file_paths_str = [p for p in row_data[2:] if p] # File paths start from column 3 (index 2)

            if not (0 <= keep_index < len(file_paths_str)):
                tqdm.write(f"Skipping row {i}: 'Keep' index {keep_index+1} is out of bounds for {len(file_paths_str)} files.")
                continue

            file_to_keep_rel = file_paths_str[keep_index]
            file_to_keep_abs = root_dir / file_to_keep_rel

            if copy_destination:
                # Copy Mode: Only the designated keeper is copied
                if file_to_keep_abs.exists():
                    tasks.append(("copy", file_to_keep_abs, copy_destination / file_to_keep_rel, i))
                else:
                    tqdm.write(f"Skipping row {i}: File to copy '{file_to_keep_rel}' does not exist.")
            else:
                # Deletion/Trash Mode: All files except the keeper are processed
                for idx, path_str in enumerate(file_paths_str):
                    if idx == keep_index:
                        continue # This is the file to keep, so skip it

                    file_to_process = root_dir / path_str
                    if file_to_process.exists():
                        tasks.append(("delete", file_to_process, None, i))
                    else:
                        tqdm.write(f"Skipping row {i}: File to delete '{path_str}' does not exist.")

        except (ValueError, TypeError, IndexError) as e:
            msg = f"ERROR on row {i}: Unexpected error processing row. Reason: {e}"
            tqdm.write(msg)
            # We don't add to failed_actions here, as that's for execution failures
    return tasks, summary


def process_report_actions(report_path: Path, root_dir: Path, dry_run: bool, use_trash: bool, copy_destination: Path | None, skip_existing: bool):
    """Reads a report file and performs actions (copy, delete, trash) based on flags."""
    if not report_path.is_file():
        print(f"Error: Report file not found at '{report_path}'")
        return

    # Find the correct sheet, preferring the new name but falling back to the old one.
    sheet_name = "Files to Process"
    if sheet_name not in workbook.sheetnames:
        if "Duplicates" in workbook.sheetnames:
            sheet_name = "Duplicates"
            print(f"Info: Sheet 'Files to Process' not found. Using existing sheet 'Duplicates'.")
        else:
            print(f"Error: Could not find a 'Files to Process' or 'Duplicates' sheet in the report.")
            return
    sheet = workbook[sheet_name]
    print("--- Processing Duplicate Report ---")
    if dry_run:
        print("--- DRY RUN MODE: No files will be changed. ---")
    elif copy_destination:
        print(f"--- COPY MODE: Keeping selected files to '{copy_destination}' ---")
    elif use_trash:
        print("--- TRASH MODE: Files will be moved to Recycle Bin/Trash ---")
    else: # This implies --execute
        print("--- DELETE MODE: Files will be permanently deleted ---")

    # --- Read Tasks from Report ---
    if report_path.suffix.lower() == '.json':
        tasks, summary_info = _read_tasks_from_json(report_path, root_dir, copy_destination)
    elif report_path.suffix.lower() == '.xlsx':
        tasks, summary_info = _read_tasks_from_xlsx(report_path, root_dir, copy_destination)
    else:
        print(f"Error: Unsupported report format '{report_path.suffix}'. Please use .json or .xlsx.")
        return

    if not tasks:
        print("No actions to perform based on the report.")
        return

    # --- Confirmation Prompt for Copy Mode ---
    if copy_destination and not dry_run:
        total_files_to_copy = summary_info.get("total_files_to_keep")
        total_size_to_copy_gb = summary_info.get("total_size_to_keep_gb")
        if total_files_to_copy is not None and total_size_to_copy_gb is not None:
            confirm = input(f"You are about to copy {total_files_to_copy} files totaling {total_size_to_copy_gb} GB to '{copy_destination}'. Proceed? (y/n): ").lower()
            if confirm != 'y':
                print("Copy operation cancelled by user.")
                return
        else:
            print("Warning: Could not read summary from report. Proceeding without confirmation.")

    # --- Main Processing Loop ---
    total_processed_files = 0 # Files successfully copied/deleted/trashed
    failed_actions = [] # List to store details of failed operations

    # Execute tasks with progress bar
    if not dry_run:
        with ThreadPoolExecutor(max_workers=os.cpu_count() - 2) as executor:
            futures = []
            for task_type, source_path, dest_path, original_row_index in tasks:
                if task_type == "copy":
                    futures.append(executor.submit(
                        _perform_single_copy,
                        source_path,
                        dest_path,
                        original_row_index,
                        root_dir,
                        skip_existing
                    ))
                elif task_type == "delete":
                    futures.append(executor.submit(
                        _perform_single_delete,
                        source_path,
                        original_row_index,
                        root_dir,
                        use_trash
                    ))
            
            for future in tqdm(as_completed(futures), total=len(futures), desc="Processing files"):
                result = future.result()
                if result: # If result is not None, it's a failed action
                    failed_actions.append(result)
                else:
                    total_processed_files += 1
    elif dry_run: # Dry run mode, just print what would happen
        for task_type, source_path, dest_path, original_row_index in tasks:
            if task_type == "copy":
                print(f"  - Would copy: '{source_path.relative_to(root_dir)}' to '{dest_path}'") # Fixed here
            elif task_type == "delete":
                action = "Move to trash" if use_trash else "Delete"
                print(f"  - Would {action}: '{source_path.relative_to(root_dir)}'")

    # If we used tqdm, the cursor is on the same line. Add a newline.
    if not dry_run:
        print()

    print("--- Action Summary ---")
    if dry_run:
        print("Dry run complete. No files were changed.")
    elif copy_destination:
        print(f"Successfully copied {total_processed_files} files to '{copy_destination}'.")
        _update_report_with_failed_actions(report_path, failed_actions)
    elif use_trash:
        print(f"Successfully moved {total_processed_files} files to trash.")
        _update_report_with_failed_actions(report_path, failed_actions)
    else:
        print(f"Successfully permanently deleted {total_processed_files} files.")
        _update_report_with_failed_actions(report_path, failed_actions)

    if failed_actions:
        print(f"Warning: {len(failed_actions)} actions failed. Check the 'Summary' sheet in the report for details.")


def _perform_single_copy(source_path: Path, dest_base_path: Path, original_row_index: int, root_dir: Path, skip_existing: bool) -> dict | None:
    """Performs a single file copy, handling conflicts and errors."""
    try:
        if not source_path.exists():
            return {"Source": str(source_path.relative_to(root_dir)), "Destination": "N/A", "Error": "Source file not found."}
        if source_path.stat().st_size == 0:
            return {"Source": str(source_path.relative_to(root_dir)), "Destination": "N/A", "Error": "Skipped: 0-byte file."}

        if skip_existing and dest_base_path.exists():
            return None # Skip if the file exists and the flag is set
        dest_path = dest_base_path
        counter = 0
        while dest_path.exists():
            if files_are_identical(source_path, dest_path):
                tqdm.write(f"Skipped identical: '{source_path.relative_to(root_dir)}'")
                return None  # Skip, no error
            else:
                counter += 1
                # Append counter before extension
                dest_path = dest_base_path.with_stem(f"{dest_base_path.stem}_{counter}")

        dest_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source_path, dest_path)
        return None # Success
    except Exception as e:
        return {"Source": str(source_path.relative_to(root_dir)), "Destination": str(dest_path.relative_to(root_dir)) if 'dest_path' in locals() else "N/A", "Error": str(e)}


def _perform_single_delete(file_path: Path, original_row_index: int, root_dir: Path, use_trash: bool) -> dict | None:
    """Performs a single file deletion or move to trash."""
    try:
        if use_trash:
            send2trash.send2trash(file_path)
        else:
            os.remove(file_path)
        return None # Success
    except Exception as e:
        return {"Source": str(file_path.relative_to(root_dir)), "Destination": "N/A", "Error": str(e)}


def _update_report_with_failed_actions(report_path: Path, failed_actions: list[dict]):
    """Adds a section for failed actions to the 'Summary' sheet of the report."""
    if not failed_actions:
        return

    try:
        workbook = openpyxl.load_workbook(report_path)
        if "Summary" not in workbook.sheetnames:
            summary_sheet = workbook.create_sheet("Summary")
        else:
            summary_sheet = workbook["Summary"]
        
        # Find the next empty row
        next_row = summary_sheet.max_row + 2 # +2 for a blank row after previous content

        summary_sheet.cell(row=next_row, column=1, value="Failed Actions Summary").font = openpyxl.styles.Font(bold=True)
        next_row += 1
        summary_sheet.cell(row=next_row, column=1, value="Source File").font = openpyxl.styles.Font(bold=True)
        summary_sheet.cell(row=next_row, column=2, value="Destination File").font = openpyxl.styles.Font(bold=True)
        summary_sheet.cell(row=next_row, column=3, value="Error").font = openpyxl.styles.Font(bold=True)
        next_row += 1

        for action in failed_actions:
            summary_sheet.cell(row=next_row, column=1, value=action.get("Source", "N/A"))
            summary_sheet.cell(row=next_row, column=2, value=action.get("Destination", "N/A"))
            summary_sheet.cell(row=next_row, column=3, value=action.get("Error", "N/A"))
            next_row += 1

        # Auto-fit columns for the new section
        for col_idx in range(1, 4):
            max_len = 0
            for row_idx in range(1, summary_sheet.max_row + 1):
                cell_value = summary_sheet.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_len = max(max_len, len(str(cell_value)))
            summary_sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_len + 2

        workbook.save(report_path)
    except Exception as e:
        print(f"ERROR: Could not update report with failed actions: {e}")


def main():
    parser = argparse.ArgumentParser(formatter_class=argparse.RawTextHelpFormatter,
        description="Processes duplicate files based on an XLSX report found within the target folder. "
                    "Choose an action: --execute (permanent delete), --trash (move to recycle bin), or --copy (copy all kept files)."
    )
    parser.add_argument(
        "folder",
        help="The root folder that was scanned, which should contain the report file.",
    )
    parser.add_argument(
        "-r",
        "--report-name",
        default=DEFAULT_REPORT_FILE,
        help=f"Name of the report file to use. Defaults to '{DEFAULT_REPORT_FILE}'.",
    )

    # Action arguments
    parser.add_argument(
        "--execute",
        action="store_true",
        help="Actually delete the files (permanent deletion). Default is a dry run.",
    )
    parser.add_argument(
        "--trash",
        action="store_true",
        help="Move deleted files to the system's trash/recycle bin instead of permanent deletion.",
    )
    parser.add_argument(
        "--copy",
        type=Path,
        help="Copy ALL kept files (unique and chosen duplicates) to this absolute destination folder, preserving relative structure. No files will be deleted.",
    )
    parser.add_argument(
        "--skip-existing",
        action="store_true",
        help="When using --copy, skip any file that already exists in the destination, regardless of content.",
    )
    args = parser.parse_args()

    root_directory = Path(args.folder).resolve()
    report_file = root_directory / args.report_name

    if not root_directory.is_dir():
        print(f"Error: Target folder not found at '{root_directory}'")
        return

    # Determine the action mode based on flags
    # If --copy is present, it's a copy operation.
    # Otherwise, it's a delete/trash operation (with dry_run as default if no --execute or --trash)
    if args.copy:
        dry_run = False # Copy implies action, unless explicitly combined with --dry-run (not implemented here)
    else:
        dry_run = not args.execute and not args.trash

    use_trash = args.trash
    copy_destination = args.copy
    skip_existing = args.skip_existing

    # Validate copy_destination if provided
    if copy_destination:
        if not copy_destination.is_absolute():
            print(f"Error: The --copy destination '{copy_destination}' must be an absolute path.")
            return
        # Ensure the directory exists, create if not (mkdir is handled inside _perform_single_copy)
        # No need to print "Created copy destination folder" here, as it might be created for subdirs

    # Check for conflicting actions
    if sum([bool(args.execute), bool(args.trash), bool(args.copy)]) > 1:
        parser.error("Please choose only one action: --execute, --trash, or --copy.")
    
    if skip_existing and not copy_destination:
        parser.error("--skip-existing can only be used with the --copy action.")

    process_report_actions(report_file, root_directory, dry_run, use_trash, copy_destination, skip_existing)


if __name__ == "__main__":
    main()