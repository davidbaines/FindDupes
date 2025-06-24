import argparse
import os
import re
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from itertools import combinations
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from tqdm import tqdm

DEFAULT_REPORT_FILE = "duplicates_report.json"

CHUNK_SIZE = 1024  # 1KB as requested


def get_file_stats(file_path: Path) -> tuple[Path, int, float] | None:
    """Returns (path, size, mtime) tuple or None if file is inaccessible/empty."""
    try:
        stats = file_path.stat()
        size = stats.st_size
        if size > 0:
            return file_path, size, stats.st_mtime
    except OSError:
        return None
    return None


def get_edge_chunks(file_path: Path, size: int) -> bytes | None:
    """Reads the first and last CHUNK_SIZE bytes of a file."""
    if size < CHUNK_SIZE * 2:
        try:
            return file_path.read_bytes()
        except OSError:
            return None
    else:
        try:
            with open(file_path, "rb") as f:
                first_chunk = f.read(CHUNK_SIZE)
                f.seek(-CHUNK_SIZE, os.SEEK_END)
                last_chunk = f.read(CHUNK_SIZE)
                return first_chunk + last_chunk
        except OSError:
            return None


def files_are_identical(file1: Path, file2: Path) -> bool:
    """Performs a full byte-by-byte comparison of two files."""
    buffer_size = 65536  # 64KB buffer
    try:
        with open(file1, "rb") as f1, open(file2, "rb") as f2:
            while True:
                b1 = f1.read(buffer_size)
                b2 = f2.read(buffer_size)
                if b1 != b2:
                    return False
                if not b1:
                    return True
    except OSError:
        return False


def get_filename_score(file_path: Path) -> int:
    """
    Assigns a score to a filename. Lower is better.
    Prefers shorter paths and names without copy indicators.
    """
    filename = file_path.name.lower()
    score = 0
    # 1. Penalize common copy indicators
    if re.search(r"\(copy\)|[-_ ]copy", filename):
        score += 10
    if re.search(r"\(\d+\)", filename):
        score += 5

    # 2. Prefer shorter paths (often indicates a more "root" or original location)
    # Using path depth (number of parent directories) is more robust than string length.
    score += len(file_path.parts) - 1

    return score


def guess_keeper(group: list[Path]) -> int:
    """
    Guesses which file to keep from a group.
    Returns the 1-based index of the best candidate.
    """
    if not group:
        return 0

    scores = [(get_filename_score(p), p) for p in group]
    scores.sort()

    best_path = scores[0][1]
    return group.index(best_path) + 1


def find_duplicates(root_dir: Path) -> tuple[list[list[Path]], list[tuple[Path, int, float]]]:
    """Finds duplicate files using a multi-stage, multithreaded approach."""
    files_by_size = defaultdict(list)
    all_paths = [p for p in root_dir.rglob("*")  if p.is_file()]
    print(f"Found {len(all_paths)} files to scan.")
    
    all_files_metadata = []

    print("Phase 1: Grouping files by size...")
    with ThreadPoolExecutor() as executor:
        with tqdm(total=len(all_paths), desc="Scanning files") as pbar:
            futures = [executor.submit(get_file_stats, path) for path in all_paths]
            for future in as_completed(futures):
                result = future.result()
                if result:
                    path, size, _ = result
                    files_by_size[size].append(path)
                    all_files_metadata.append(result)
                pbar.update(1)

    potential_dupes_by_size = {
        s: p for s, p in files_by_size.items() if len(p) > 1
    }
    print(f"\nFound {len(potential_dupes_by_size)} groups of files with the same size.")

    print("\nPhase 2: Comparing file edges (first and last 1KB)...")
    likely_dupes = []
    with ThreadPoolExecutor() as executor:
        with tqdm(total=len(potential_dupes_by_size), desc="Checking edges") as pbar:
            for size, files in potential_dupes_by_size.items():
                chunks_by_hash = defaultdict(list)
                # Submit all chunk reads for the current size group
                future_to_path = {
                    executor.submit(get_edge_chunks, path, size): path for path in files
                }
                for future in as_completed(future_to_path):
                    path = future_to_path[future]
                    chunks = future.result()
                    if chunks:
                        chunks_by_hash[chunks].append(path)

                for group in chunks_by_hash.values():
                    if len(group) > 1:
                        likely_dupes.append(group)
                pbar.update(1)

    print(f"\nFound {len(likely_dupes)} groups of files with matching edges.")

    print("\nPhase 3: Final byte-by-byte verification...")
    confirmed_duplicates = []
    for group in tqdm(likely_dupes, desc="Verifying duplicates"):
        checked_files = set()
        for file1, file2 in combinations(group, 2):
            # This check ensures we don't form multiple pairs from the same group
            if file1 in checked_files and file2 in checked_files:
                continue

            if files_are_identical(file1, file2):
                # Find if either file is already in a confirmed set
                found_set = None
                for s in confirmed_duplicates:
                    if file1 in s or file2 in s:
                        found_set = s
                        break
                if found_set is not None:
                    found_set.add(file1)
                    found_set.add(file2)
                else:
                    confirmed_duplicates.append({file1, file2})
                checked_files.add(file1)
                checked_files.add(file2)

    # Convert sets to sorted lists for consistent output
    return [sorted(list(s)) for s in confirmed_duplicates], all_files_metadata


def create_json_report(
    all_files_metadata: list[tuple[Path, int, float]],
    confirmed_duplicates: list[list[Path]],
    root_dir: Path,
    output_file: Path,
):
    """Creates a fast, efficient JSON report of all files."""
    report_data = {"files_to_process": []}
    kept_files_info = []

    # Identify all files that are part of a duplicate group
    duplicate_file_paths_set = {f_path for group in confirmed_duplicates for f_path in group}

    # Map of file path to its group
    file_to_group_map = {f_path: group for group in confirmed_duplicates for f_path in group}

    processed_groups = set()
    for file_path, size, _ in tqdm(all_files_metadata, desc="Generating JSON report"):
        if file_path in duplicate_file_paths_set:
            group = file_to_group_map[file_path]
            if id(group) not in processed_groups:
                keeper_idx_1_based = guess_keeper(group)
                kept_files_info.append((group[keeper_idx_1_based - 1], size))
                report_data["files_to_process"].append({
                    "type": "duplicate_group",
                    "size_bytes": size,
                    "files": [str(p.relative_to(root_dir)) for p in group],
                    "keep_index": keeper_idx_1_based - 1,  # 0-based index
                })
                processed_groups.add(id(group))
        else:  # Unique file
            kept_files_info.append((file_path, size))
            report_data["files_to_process"].append({
                "type": "unique_file",
                "size_bytes": size,
                "file": str(file_path.relative_to(root_dir)),
            })

    total_kept_size_bytes = sum(s for _, s in kept_files_info)
    report_data["summary"] = {
        "total_files_to_keep": len(kept_files_info),
        "total_size_to_keep_bytes": total_kept_size_bytes,
        "total_size_to_keep_gb": f"{total_kept_size_bytes / (1024**3):.2f}",
    }

    with open(output_file, "w") as f:
        json.dump(report_data, f, indent=2)

def create_xlsx_report( # Modified signature
    all_files_metadata: list[tuple[Path, int, float]],
    confirmed_duplicates: list[list[Path]],
    root_dir: Path,
    output_file: Path,
):
    workbook = openpyxl.Workbook()
    sheet = workbook.active # Get the default sheet
    sheet.title = "Files to Process" # Renamed for clarity

    # Determine max files in any group (duplicate or unique) for header
    max_files_in_group = 0
    if confirmed_duplicates:
        max_files_in_group = max(len(s) for s in confirmed_duplicates)
    # Unique files are treated as groups of 1
    max_files_in_group = max(max_files_in_group, 1)

    headers = ["Keep", "Size (Bytes)"] + [
        f"File {i+1}" for i in range(max_files_in_group)
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    # Identify all files that are part of a duplicate group
    duplicate_file_paths_set = set()
    for group in confirmed_duplicates:
        for f_path in group:
            duplicate_file_paths_set.add(f_path)

    # Map of file path to its group (for duplicates) or None (for unique)
    file_to_group_map = {}
    for group in confirmed_duplicates:
        for f_path in group:
            file_to_group_map[f_path] = group

    processed_groups = set() # To avoid processing the same duplicate group multiple times
    kept_files_info = [] # List of (Path, size) for all files that will be kept

    for file_path, size, _ in tqdm(all_files_metadata, desc="Populating report sheet"):
        if file_path in duplicate_file_paths_set:
            group = file_to_group_map[file_path]
            # Only process the group once
            if id(group) not in processed_groups: # Use id() for unique group identification
                keeper_idx = guess_keeper(group)
                kept_files_info.append((group[keeper_idx-1], size)) # Add keeper to kept_files_info

                row = [keeper_idx, size]
                relative_paths = [str(p.relative_to(root_dir)) for p in group]
                row.extend(relative_paths)
                sheet.append(row)
                processed_groups.add(id(group))
        else: # Unique file
            kept_files_info.append((file_path, size)) # Add unique file to kept_files_info

            row = [1, size, str(file_path.relative_to(root_dir))] # Unique files are group of 1
            sheet.append(row)

    # Auto-fit columns
    for column_cells in sheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

    # --- New Summary Tab ---
    summary_sheet = workbook.create_sheet("Summary")
    total_kept_files_count = len(kept_files_info)
    total_kept_files_size_bytes = sum(size for _, size in kept_files_info)
    total_kept_files_size_gb = total_kept_files_size_bytes / (1024**3)

    summary_sheet.append(["Metric", "Value"])
    summary_sheet.append(["Total Files to Keep", total_kept_files_count])
    summary_sheet.append(["Total Size to Keep (Bytes)", total_kept_files_size_bytes])
    summary_sheet.append(["Total Size to Keep (GB)", f"{total_kept_files_size_gb:.2f}"])

    # Auto-fit columns for summary sheet
    for column_cells in summary_sheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        summary_sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Remove the default "Sheet" if it exists and is empty
    if "Sheet" in workbook.sheetnames and workbook["Sheet"].max_row == 0:
        workbook.remove(workbook["Sheet"])

    workbook.save(output_file)
    print(f"\nSuccessfully created report: {output_file}")
    print(f"Summary: {total_kept_files_count} files to keep, totaling {total_kept_files_size_gb:.2f} GB.")


def main():
    parser = argparse.ArgumentParser(
        description="Find duplicate files and generate a report for processing."
    )
    parser.add_argument("folder", help="The root folder to scan for duplicates.")
    parser.add_argument(
        "-o",
        "--output",
        default=f"{DEFAULT_REPORT_FILE}",
        help=f"Name of the output report file. Default is {DEFAULT_REPORT_FILE}.",
    )
    parser.add_argument(
        "--format",
        choices=["json", "xlsx"],
        default="json",
        help="The output format for the report. 'json' is much faster. Default is 'json'."
    )
    args = parser.parse_args()

    root_directory = Path(args.folder).resolve()
    output_path = root_directory / args.output

    if not root_directory.is_dir():
        print(f"Error: Directory not found at '{root_directory}'")
        return

    duplicates, all_files_metadata = find_duplicates(root_directory) # find_duplicates now returns all_files_metadata
    
    print(f"\nCreating {args.format.upper()} report...")
    if args.format == "json":
        create_json_report(all_files_metadata, duplicates, root_directory, output_path)
    else: # xlsx
        create_xlsx_report(all_files_metadata, duplicates, root_directory, output_path)


if __name__ == "__main__":
    main()