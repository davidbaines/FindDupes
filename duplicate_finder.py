import argparse
import json  # For caching
import os
import random  # For random sampling
import shutil
from collections import defaultdict
from datetime import datetime
from itertools import combinations
from pathlib import Path

import openpyxl
import xxhash
from tqdm import tqdm

CHUNK_SIZE = 4096  # Read files in 4KB chunks


def get_partial_hash(file: Path) -> str:
    """
    Calculates a hash from the first and last chunk of a file.
    This provides a more robust check than just the first chunk alone.
    """
    hasher = xxhash.xxh64()
    try:
        size = file.stat().st_size
        if size < CHUNK_SIZE * 2:
            # For small files, just hash the whole thing to avoid read overlap
            with file.open("rb") as f:
                hasher.update(f.read())
        else:
            with file.open("rb") as f:
                # Hash the first chunk
                hasher.update(f.read(CHUNK_SIZE))
                # Seek to the end and hash the last chunk
                f.seek(-CHUNK_SIZE, os.SEEK_END)
                hasher.update(f.read(CHUNK_SIZE))

        return hasher.hexdigest()
    except (IOError, OSError):
        # File might be unreadable (permissions) or a broken symlink
        return ""


def get_full_hash(file: Path) -> str:
    """Calculates a hash for the entire content of a file."""
    hasher = xxhash.xxh64()
    try:
        with file.open("rb") as f:
            while chunk := f.read(CHUNK_SIZE):
                hasher.update(chunk)
        return hasher.hexdigest()
    except (IOError, OSError):
        return ""


def save_cache(
    confirmed_duplicates: list[list[Path]],
    all_files_metadata: list[tuple[Path, int, float]],  # (file_path, size, mtime)
    full_hashes_map: dict[str, list[Path]],
    cache_path: Path,
):
    """Saves scan results to a JSON cache file."""
    serializable_confirmed_duplicates = [
        [str(f) for f in group] for group in confirmed_duplicates
    ]
    serializable_all_files_metadata = [[str(f), s, m] for f, s, m in all_files_metadata]
    serializable_full_hashes_map = {
        h: [str(f) for f in files] for h, files in full_hashes_map.items()
    }

    cache_data = {
        "confirmed_duplicates": serializable_confirmed_duplicates,
        "all_files_metadata": serializable_all_files_metadata,
        "full_hashes_map": serializable_full_hashes_map,
    }
    with open(cache_path, "w") as f:
        json.dump(cache_data, f, indent=4)
    print(f"Scan results cached to {cache_path}")


def load_cache(
    cache_path: Path,
) -> tuple[list[list[Path]], list[tuple[Path, int, float]], dict[str, list[Path]]]:
    """Loads scan results from a JSON cache file."""
    with open(cache_path, "r") as f:
        cache_data = json.load(f)

    confirmed_duplicates = [
        [Path(f) for f in group] for group in cache_data["confirmed_duplicates"]
    ]
    all_files_metadata = [
        (Path(f), s, m) for f, s, m in cache_data["all_files_metadata"]
    ]
    full_hashes_map = {
        h: [Path(f) for f in files]
        for h, files in cache_data["full_hashes_map"].items()
    }

    return confirmed_duplicates, all_files_metadata, full_hashes_map


def find_duplicates(
    root_dir: Path,
    force_rescan: bool = False,  # New argument
) -> tuple[list[list[Path]], list[Path], dict[str, list[Path]]]:
    """
    Finds duplicate files in a directory using a multi-stage approach.
    Supports caching of previous scan results for faster subsequent runs.

    Args:
        root_dir: The directory to scan for duplicates.
        force_rescan: If True, forces a full scan, ignoring any cache.

    Returns:
        A tuple containing:
        - A list of lists, where each inner list contains paths to confirmed duplicate files.
        - A list of all file paths scanned (derived from metadata for consistency).
        - A dictionary mapping a full file hash to a list of file paths for all likely duplicates.
    """
    if not root_dir.is_dir():
        raise ValueError(f"'{root_dir}' is not a valid directory.")

    cache_path = root_dir / "finddupes_cache.json"
    CONSISTENCY_CHECK_PERCENTAGE = 0.10  # 10%

    if cache_path.exists() and not force_rescan:
        print(f"Cache file found at {cache_path}. Attempting to load and verify...")
        try:
            (
                cached_duplicates,
                cached_all_files_metadata,
                cached_full_hashes_map,
            ) = load_cache(cache_path)

            # Perform consistency check
            print(
                f"Performing consistency check on {len(cached_all_files_metadata)} files..."
            )
            num_to_check = max(
                1, int(len(cached_all_files_metadata) * CONSISTENCY_CHECK_PERCENTAGE)
            )

            # Randomly sample files for consistency check
            files_to_check_sample = random.sample(
                cached_all_files_metadata, num_to_check
            )

            is_consistent = True
            for file_path, cached_size, cached_mtime in tqdm(
                files_to_check_sample, desc="Verifying cache consistency"
            ):
                try:
                    current_stat = file_path.stat()
                    # Allow 1 second tolerance for mtime
                    if (
                        not file_path.exists()
                        or current_stat.st_size != cached_size
                        or abs(current_stat.st_mtime - cached_mtime) > 1
                    ):
                        is_consistent = False
                        tqdm.write(
                            f"Cache inconsistency detected for {file_path}. Metadata changed or file missing."
                        )
                        break
                except OSError:
                    is_consistent = False
                    tqdm.write(f"Cache inconsistency: Could not access {file_path}.")
                    break

            if is_consistent:
                print("Cache is consistent. Using cached data.")
                # Reconstruct all_files (just paths) from cached_all_files_metadata
                all_files_paths = [f for f, _, _ in cached_all_files_metadata]
                return cached_duplicates, all_files_paths, cached_full_hashes_map
            else:
                print("Cache is inconsistent. Performing full scan.")
                # Clean up potentially corrupted/outdated cache file
                cache_path.unlink(missing_ok=True)

        except (json.JSONDecodeError, KeyError, IndexError) as e:
            print(f"Error loading or parsing cache file: {e}. Performing full scan.")
            # Clean up potentially corrupted cache file
            cache_path.unlink(missing_ok=True)

    # --- Full Scan Logic (if cache not used or inconsistent) ---
    print("Stage 1: Finding files with the same size...")
    # Stage 1: Group files by size
    files_by_size = defaultdict(list)

    # Collect all files with their metadata for the full scan and for caching
    all_files_for_scan_and_cache = []
    raw_files_list = [
        file for file in root_dir.rglob("*") if file.is_file() and not file.is_symlink()
    ]
    print(f"Discovered {len(raw_files_list)} files. Now grouping them by size...")

    for file in tqdm(raw_files_list, desc="Grouping by size"):
        try:
            stat = file.stat()
            size = stat.st_size
            mtime = stat.st_mtime
            # Ignore empty files for this example, or handle them separately
            if size > 0:
                files_by_size[size].append(file)
                all_files_for_scan_and_cache.append(
                    (file, size, mtime)
                )  # Store metadata for caching
        except OSError:
            # Use tqdm.write to avoid messing up the progress bar
            tqdm.write(f"Warning: Could not access {file}")
            continue

    # Filter out unique sizes
    potential_duplicates = {
        size: files for size, files in files_by_size.items() if len(files) > 1
    }
    print(
        f"Found {len(potential_duplicates)} groups of files with potential duplicates."
    )

    print("\nStage 2: Filtering by partial hash of the first 4KB...")
    # Stage 2: Group potential duplicates by a partial hash
    potential_duplicates_by_partial_hash = defaultdict(list)
    for size in tqdm(potential_duplicates, desc="Filtering by partial hash"):
        for file in potential_duplicates[size]:
            partial_hash = get_partial_hash(file)
            if partial_hash:
                # Create a unique key from size and partial hash
                key = f"{size}:{partial_hash}"
                potential_duplicates_by_partial_hash[key].append(file)

    # Filter out unique partial hashes
    likely_duplicates = {
        key: files
        for key, files in potential_duplicates_by_partial_hash.items()
        if len(files) > 1
    }
    print(
        f"Found {len(likely_duplicates)} groups of files that are very likely duplicates."
    )

    print("\nStage 3: Verifying with full file hash...")
    # Stage 3: Group all likely duplicates by their full hash
    full_hashes_map = defaultdict(list)
    # Flatten the list of files to check to avoid re-hashing the same file
    files_to_hash = {file for files in likely_duplicates.values() for file in files}

    for file in tqdm(files_to_hash, desc="Verifying with full hash"):
        full_hash = get_full_hash(file)
        if full_hash:
            full_hashes_map[full_hash].append(file)

    # Filter for hashes that have more than one file, these are the duplicates
    confirmed_duplicates = [
        files for files in full_hashes_map.values() if len(files) > 1
    ]

    # Save cache before returning
    save_cache(
        confirmed_duplicates, all_files_for_scan_and_cache, full_hashes_map, cache_path
    )

    # Return all_files as just paths, as expected by other functions
    all_files_paths_for_return = [f for f, _, _ in all_files_for_scan_and_cache]
    return confirmed_duplicates, all_files_paths_for_return, full_hashes_map


def analyze_folder_duplicates(
    all_files: list[Path],
    full_hashes_map: dict[str, list[Path]],
) -> tuple[list[dict], dict]:  # Changed return type hint to reflect the actual return
    """
    Analyzes folder relationships based on the full content of all files within them.
    It identifies folders where the set of ALL file hashes is an exact match
    or a subset of another folder's.
    This function reuses pre-calculated hashes for duplicate files and hashes unique files.

    Args:
        all_files: A list of all files found in the scan.
        full_hashes_map: A map of file hashes to file paths.

    Returns:
        A tuple containing:
        - A list of dictionaries, each describing a redundant folder relationship.
        - A dictionary containing stats for each folder.
    """
    print("\nPhase 2: Analyzing folder relationships based on duplicate files...")

    # --- 1. Create content signatures for every folder (all files, not just duplicates) ---
    folder_content_signatures = defaultdict(set)
    all_file_hashes = {
        file: full_hash
        for full_hash, files_list in full_hashes_map.items()
        for file in files_list
    }

    for file in tqdm(all_files, desc="Building folder content signatures"):
        try:
            if file not in all_file_hashes:
                file_hash = get_full_hash(file)
                if file_hash:
                    all_file_hashes[file] = file_hash
            else:
                file_hash = all_file_hashes[file]  # Use already known hash

            if file_hash:
                folder_content_signatures[file.parent].add(file_hash)
        except OSError:
            continue

    # --- 2. Calculate overall stats (including last modified time) for all folders ---
    folder_stats = defaultdict(lambda: {"size": 0, "files": 0, "last_modified": 0.0})
    for file in tqdm(all_files, desc="Gathering folder stats"):
        try:
            stat = file.stat()
            folder = file.parent
            folder_stats[folder]["files"] += 1
            folder_stats[folder]["size"] += stat.st_size
            folder_stats[folder]["last_modified"] = max(
                folder_stats[folder]["last_modified"], stat.st_mtime
            )
        except OSError:
            continue

    # --- 3. Compare folder content signatures to find relationships ---
    folder_relationships = []
    # Use combinations to compare each pair of folders only once
    folders_to_compare = list(folder_content_signatures.keys())
    # The number of combinations can be large, so add a progress bar
    num_combinations = (
        len(folders_to_compare) * (len(folders_to_compare) - 1) // 2
        if len(folders_to_compare) > 1
        else 0
    )
    for folder_a, folder_b in tqdm(
        combinations(folders_to_compare, 2),
        total=num_combinations,
        desc="Comparing folder contents",
    ):
        hashes_a = folder_content_signatures[folder_a]
        hashes_b = folder_content_signatures[folder_b]

        # Only consider relationships if both folders have content
        if not hashes_a or not hashes_b:
            continue

        if hashes_a == hashes_b:
            # Exact duplicates (all files are identical)
            folder_relationships.append(
                {"type": "Exact Duplicate", "folder_a": folder_a, "folder_b": folder_b}
            )
        elif hashes_a.issubset(hashes_b):
            # Folder A is a subset of Folder B (all files in A are in B, B has more)
            folder_relationships.append(
                {"type": "Subset", "subset": folder_a, "superset": folder_b}
            )
        elif hashes_b.issubset(hashes_a):
            # Folder B is a subset of Folder A (all files in B are in A, A has more)
            folder_relationships.append(
                {"type": "Subset", "subset": folder_b, "superset": folder_a}
            )

    return folder_relationships, folder_stats


def process_interactive_file_deletions(duplicate_sets: list[list[Path]]):
    """Handles the interactive deletion process for duplicate files."""
    print("\n--- Interactive File Deletion Mode ---")
    if not duplicate_sets:
        print("No duplicate files found to process.")
        return

    # Sort sets by file size, largest first, for maximum impact
    try:
        duplicate_sets.sort(key=lambda group: group[0].stat().st_size, reverse=True)
    except OSError:
        print(
            "Warning: Could not stat a file to sort sets by size. Processing in default order."
        )

    for group in duplicate_sets:
        # First, filter the group to only include files that still exist on disk.
        # This handles cases where a file was deleted as part of a previous folder operation.
        current_group = [file for file in group if file.exists()]
        current_group.sort()  # Sort by path for predictable order

        # If after filtering, we have fewer than 2 files, it's no longer a duplicate set.
        if len(current_group) < 2:
            continue

        try:
            size_mb = current_group[0].stat().st_size / (1024 * 1024)
            print("\n" + "=" * 80)
            print(f"Duplicate set found (Size: {size_mb:.2f} MB):")
        except OSError:
            print("\n" + "=" * 80)
            print("Duplicate set found (size unknown):")

        file_options = {}
        for i, file in enumerate(current_group, 1):
            file_options[str(i)] = file
            try:
                last_modified = datetime.fromtimestamp(file.stat().st_mtime).strftime(
                    "%Y-%m-%d %H:%M:%S"
                )
                print(f"[{i}] Path: {file}\n    Last Modified: {last_modified}")
            except OSError:
                print(f"[{i}] Path: {file}\n    Last Modified: <Could not read>")

        # Loop for this specific group until an action is taken (delete/skip)
        while True:
            prompt = f"Which file to KEEP? [{'/'.join(file_options.keys())}] or (F)older mode or (S)kip: "
            choice = input(prompt)

            if choice.upper() == "S":
                print("Skipping.")
                break  # Action complete for this group, move to the next.

            elif choice.upper() == "F":
                parent_folders = sorted(list({file.parent for file in current_group}))
                if len(parent_folders) <= 1:
                    print(
                        "All duplicates are in the same folder. Cannot use folder mode."
                    )
                    continue  # Re-ask the question for this group.

                print("\n--- Folder Deletion Mode ---")
                folder_options = {
                    str(i): folder for i, folder in enumerate(parent_folders, 1)
                }
                for i_str, folder in folder_options.items():
                    print(f"[{i_str}] {folder}")

                folder_choice = ""
                while folder_choice.upper() not in list(folder_options.keys()) + ["C"]:
                    folder_choice = input(
                        f"Which folder to KEEP? All other listed folders will be deleted. [{'/'.join(folder_options.keys())}] or (C)ancel: "
                    )

                if folder_choice.upper() == "C":
                    print("Cancelling folder selection.")
                    continue  # Re-ask the question for this group.

                folder_to_keep = folder_options[folder_choice]
                folders_to_delete = [f for f in parent_folders if f != folder_to_keep]

                print(
                    "\nWARNING: The following folders and ALL THEIR CONTENTS will be recursively deleted:"
                )
                for f in folders_to_delete:
                    print(f"  - {f}")

                confirm = input("Are you sure you want to proceed? (y/n): ")
                if confirm.lower() == "y":
                    for folder in folders_to_delete:
                        try:
                            print(f"Deleting folder: {folder}")
                            shutil.rmtree(folder)
                            print(f"Successfully deleted {folder}")
                        except OSError as e:
                            print(f"ERROR: Could not delete {folder}. Reason: {e}")
                    break  # Action complete for this group, move to the next.
                else:
                    print("Folder deletion cancelled.")
                    continue  # Re-ask the question for this group.

            elif choice in file_options:
                file_to_keep = file_options[choice]
                print(f"Keeping: {file_to_keep}")
                for file_to_delete in current_group:
                    if file_to_delete != file_to_keep:
                        try:
                            print(f"  - Deleting: {file_to_delete}")
                            file_to_delete.unlink()
                        except OSError as e:
                            print(
                                f"    ERROR: Could not delete {file_to_delete}. Reason: {e}"
                            )
                break  # Action complete for this group, move to the next.

    print("\n--- Interactive File Deletion Finished ---")


def process_interactive_deletions(folder_relationships: list[dict], folder_stats: dict):
    """Handles the interactive deletion process for redundant folders."""
    print("\n--- Interactive Deletion Mode ---")
    if not folder_relationships:
        print("No redundant folder relationships found to process.")
        return

    deleted_folders = set()
    for rel in folder_relationships:
        folder_a_path, folder_b_path = (
            (rel.get("folder_a"), rel.get("folder_b"))
            if rel["type"] == "Exact Duplicate"
            else (rel.get("subset"), rel.get("superset"))
        )

        # Skip if either folder has already been deleted in this session
        if folder_a_path in deleted_folders or folder_b_path in deleted_folders:
            continue

        info_a = folder_stats[folder_a_path]
        info_b = folder_stats[folder_b_path]

        # Prepare display strings
        dt_a = datetime.fromtimestamp(info_a["last_modified"]).strftime(
            "%Y-%m-%d %H:%M:%S"
        )
        dt_b = datetime.fromtimestamp(info_b["last_modified"]).strftime(
            "%Y-%m-%d %H:%M:%S"
        )
        size_a_mb = info_a["size"] / (1024 * 1024)
        size_b_mb = info_b["size"] / (1024 * 1024)

        print("\n" + "=" * 80)
        print(
            f"[A] Path: {folder_a_path}\n    Files: {info_a['files']}, Size: {size_a_mb:.2f} MB, Last Modified: {dt_a}"
        )
        print(
            f"[B] Path: {folder_b_path}\n    Files: {info_b['files']}, Size: {size_b_mb:.2f} MB, Last Modified: {dt_b}"
        )

        if rel["type"] == "Exact Duplicate":
            print("Relationship: Folders are EXACT DUPLICATES.")
        else:
            print(f"Relationship: Folder [A] is a SUBSET of folder [B].")

        # Get user input
        choice = ""
        while choice.upper() not in ["A", "B", "S"]:
            choice = input("Which folder to DELETE? [A] or [B] or (S)kip: ")

        folder_to_delete = None
        if choice.upper() == "A":
            folder_to_delete = folder_a_path
        elif choice.upper() == "B":
            folder_to_delete = folder_b_path
        else:
            print("Skipping.")
            continue

        # Perform deletion
        if folder_to_delete:
            try:
                print(f"Deleting {folder_to_delete}...")
                shutil.rmtree(folder_to_delete)
                print(f"Successfully deleted {folder_to_delete}")
                deleted_folders.add(folder_to_delete)
            except OSError as e:
                print(f"ERROR: Could not delete {folder_to_delete}. Reason: {e}")

    print("\n--- Interactive Deletion Finished ---")


def create_xlsx_report(
    duplicate_sets: list[list[Path]],
    folder_analysis: tuple[list[dict], dict],
    output_dir: Path,
):
    """
    Creates an XLSX report listing all found duplicate files with their metadata.
    Now includes tabs for folder-level analysis.

    Args:
        duplicate_sets: The list of duplicate file groups.
        folder_analysis: The dictionary with results from folder analysis.
        output_dir: The directory where the report will be saved.
    """

    workbook = openpyxl.Workbook()

    # --- Tab 1: File Duplicates ---
    sheet = workbook.create_sheet("File Duplicates", 0)

    # Create and style the header row
    header = [
        "Set ID",
        "File Path",
        "Size (Bytes)",
        "File Type",
        "Date Created",
        "Date Modified",
    ]
    sheet.append(header)
    for cell in sheet[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    # Populate the report with file data
    set_id_counter = 1
    for group in duplicate_sets:
        for file in group:
            try:
                stat = file.stat()
                created = datetime.fromtimestamp(stat.st_ctime).strftime(
                    "%Y-%m-%d %H:%M:%S"
                )
                modified = datetime.fromtimestamp(stat.st_mtime).strftime(
                    "%Y-%m-%d %H:%M:%S"
                )
                row = [
                    set_id_counter,
                    str(file),
                    stat.st_size,
                    file.suffix or "N/A",
                    created,
                    modified,
                ]
                sheet.append(row)
            except OSError as e:
                # Append a row with error info if stats can't be read
                sheet.append([set_id_counter, str(file), f"Error: {e}"])
        set_id_counter += 1

    # --- Tab 2: Redundant Folders ---
    folder_relationships, folder_info = folder_analysis
    if folder_relationships:
        sheet2 = workbook.create_sheet("Redundant Folders", 1)
        sheet2.append(
            [
                "Folder A Path",
                "Folder A Files",
                "Folder A Size (Bytes)",
                "Folder A Last Modified",
                "Relationship",
                "Folder B Path",
                "Folder B Files",
                "Folder B Size (Bytes)",
                "Folder B Last Modified",
            ]
        )

        for rel in folder_relationships:
            row = []
            if rel["type"] == "Exact Duplicate":
                folder_a, folder_b = rel["folder_a"], rel["folder_b"]
                info_a = folder_info[folder_a]
                info_b = folder_info[folder_b]
                row = [
                    str(folder_a),
                    info_a["files"],
                    info_a["size"],
                    datetime.fromtimestamp(info_a["last_modified"]),
                    "Exact Duplicate",
                    str(folder_b),
                    info_b["files"],
                    info_b["size"],
                    datetime.fromtimestamp(info_b["last_modified"]),
                ]
            elif rel["type"] == "Subset":
                subset_folder = rel["subset"]
                superset_folder = rel["superset"]
                info_subset = folder_info[subset_folder]
                info_superset = folder_info[superset_folder]
                row = [
                    str(subset_folder),
                    info_subset["files"],
                    info_subset["size"],
                    datetime.fromtimestamp(info_subset["last_modified"]),
                    "Is a Subset of",
                    str(superset_folder),
                    info_superset["files"],
                    info_superset["size"],
                    datetime.fromtimestamp(info_superset["last_modified"]),
                ]
            if row:
                sheet2.append(row)

    # Remove the default sheet created by openpyxl
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    # Auto-fit columns for all sheets
    for sheet in workbook.worksheets:
        for column_cells in sheet.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

    report_path = output_dir / "duplicates_report.xlsx"
    workbook.save(report_path)
    print(f"\nSuccessfully created duplicates report at: {report_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Find duplicate files in a specified folder and generate a report."
    )
    parser.add_argument("folder", help="The root folder to scan for duplicate files.")
    parser.add_argument(
        "--interactive-delete",
        action="store_true",
        help="After scanning, enter interactive mode to delete redundant folders.",
    )
    parser.add_argument(
        "--delete-files",
        action="store_true",
        help="After scanning, enter interactive mode to delete duplicate files.",
    )
    parser.add_argument(
        "--force-rescan",
        action="store_true",
        help="Force a full rescan, ignoring any cached data.",
    )
    args = parser.parse_args()
    target_directory = Path(args.folder)

    try:
        duplicate_sets, all_files, full_hashes_map = find_duplicates(
            target_directory, args.force_rescan
        )

        if not duplicate_sets:
            print("\nNo duplicate files found.")
        else:
            print(f"\n--- Found {len(duplicate_sets)} sets of duplicate files ---")

            if args.delete_files:
                process_interactive_file_deletions(duplicate_sets)
            elif args.interactive_delete:
                # Enter interactive folder deletion mode
                folder_analysis_results = analyze_folder_duplicates(
                    all_files, full_hashes_map
                )
                relationships, stats = folder_analysis_results
                process_interactive_deletions(relationships, stats)
            else:
                # Default behavior: generate a report
                folder_analysis_results = analyze_folder_duplicates(
                    all_files, full_hashes_map
                )
                print("Phase 3: Generating XLSX report...")
                create_xlsx_report(
                    duplicate_sets, folder_analysis_results, target_directory
                )

    except ValueError as e:
        print(f"Error: {e}")
    except KeyboardInterrupt:
        print("\nScan cancelled by user.")
