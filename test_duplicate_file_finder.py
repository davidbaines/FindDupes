import unittest
import tempfile
import shutil
import io
from contextlib import redirect_stdout
from pathlib import Path
import openpyxl

from duplicate_file_finder import ( # Changed import
    find_duplicates,
    create_xlsx_report,
    get_filename_score,
    guess_keeper,
    get_file_stats,
    get_edge_chunks,
    files_are_identical,
)


class TestCreateDuplicateReport(unittest.TestCase):
    def setUp(self):
        self.test_dir = Path(tempfile.mkdtemp())

    def tearDown(self):
        shutil.rmtree(self.test_dir)

    def create_dummy_file(self, filename, content=b""):
        filepath = self.test_dir / filename
        with open(filepath, "wb") as f:
            f.write(content)
        return filepath

    def test_get_file_stats(self):
        file1 = self.create_dummy_file("file1.txt", b"test")
        file2 = self.test_dir / "nonexistent.txt"

        self.assertEqual(get_file_stats(file1), (file1, 4, file1.stat().st_mtime))
        self.assertIsNone(get_file_stats(file2))

    def test_get_edge_chunks(self):
        file1 = self.create_dummy_file("file1.txt", b"a" * 2048)
        chunks = get_edge_chunks(file1, 2048)
        self.assertEqual(len(chunks), 2048)
        self.assertEqual(chunks[:1024], b"a" * 1024)
        self.assertEqual(chunks[1024:], b"a" * 1024)

        file2 = self.create_dummy_file("file2.txt", b"short")
        chunks = get_edge_chunks(file2, 5)
        self.assertEqual(chunks, b"short")

    def test_files_are_identical(self):
        file1 = self.create_dummy_file("file1.txt", b"identical")
        file2 = self.create_dummy_file("file2.txt", b"identical")
        file3 = self.create_dummy_file("file3.txt", b"different")

        self.assertTrue(files_are_identical(file1, file2))
        self.assertFalse(files_are_identical(file1, file3))

    def test_get_filename_score(self):
        self.assertEqual(get_filename_score(Path("file.txt")), 0)
        self.assertGreater(get_filename_score(Path("file (copy).txt")), 0)
        self.assertGreater(get_filename_score(Path("file_copy.txt")), 0)
        self.assertGreater(get_filename_score(Path("file (1).txt")), 0)

    def test_guess_keeper(self):
        files = [
            self.create_dummy_file("file (copy).txt"),
            self.create_dummy_file("file.txt"),
        ]
        self.assertEqual(guess_keeper(files), 2)  # Expect "file.txt" (index 1)

    def test_find_duplicates(self):
        # find_duplicates now returns all_files_metadata as well
        self.create_dummy_file("file1.txt", b"duplicate")
        self.create_dummy_file("file2.txt", b"duplicate")
        self.create_dummy_file("file3.txt", b"unique")

        with redirect_stdout(io.StringIO()):
            duplicates, all_files_metadata = find_duplicates(self.test_dir)
        self.assertEqual(len(duplicates), 1)
        self.assertEqual(len(duplicates[0]), 2)
        self.assertEqual(len(all_files_metadata), 3) # All 3 files should be in metadata

    def test_find_duplicates_empty_directory(self):
        with redirect_stdout(io.StringIO()):
            duplicates, all_files_metadata = find_duplicates(self.test_dir)
        self.assertEqual(len(duplicates), 0)
        self.assertEqual(len(all_files_metadata), 0)

    def test_create_xlsx_report(self):
        # Create files: 2 duplicates, 1 unique
        file1 = self.create_dummy_file("file1.txt", b"duplicate_content")
        file2 = self.create_dummy_file("file2.txt", b"duplicate_content")
        file3 = self.create_dummy_file("file3.txt", b"unique_content")

        # Simulate find_duplicates output
        # confirmed_duplicates will have one group: [file1, file2]
        confirmed_duplicates = [[file1, file2]]
        # all_files_metadata will have all three files
        all_files_metadata = [
            (file1, file1.stat().st_size, file1.stat().st_mtime),
            (file2, file2.stat().st_size, file2.stat().st_mtime),
            (file3, file3.stat().st_size, file3.stat().st_mtime),
        ]

        output_file = self.test_dir / "report.xlsx"

        # Suppress the print output from the function during the test
        with redirect_stdout(io.StringIO()):
            # create_xlsx_report now takes all_files_metadata first
            create_xlsx_report(all_files_metadata, confirmed_duplicates, self.test_dir, output_file)

        self.assertTrue(output_file.exists())

        workbook = openpyxl.load_workbook(output_file)

        # Validate "Files to Process" sheet (renamed from "Duplicates")
        self.assertIn("Files to Process", workbook.sheetnames)
        sheet = workbook["Files to Process"]
        headers = [cell.value for cell in sheet[1]]
        self.assertEqual(headers, ["Keep", "Size (Bytes)", "File 1", "File 2"]) # Max files in group is 2
        self.assertEqual(sheet.max_row, 3) # Header + 1 duplicate group row + 1 unique file row

        # Check content of "Files to Process"
        # Row 2: Duplicate group
        self.assertEqual(sheet.cell(row=2, column=1).value, 1) # Assuming file1 is guessed keeper
        self.assertEqual(sheet.cell(row=2, column=2).value, file1.stat().st_size)
        self.assertEqual(sheet.cell(row=2, column=3).value, str(file1.relative_to(self.test_dir)))
        self.assertEqual(sheet.cell(row=2, column=4).value, str(file2.relative_to(self.test_dir)))

        # Row 3: Unique file
        self.assertEqual(sheet.cell(row=3, column=1).value, 1)
        self.assertEqual(sheet.cell(row=3, column=2).value, file3.stat().st_size)
        self.assertEqual(sheet.cell(row=3, column=3).value, str(file3.relative_to(self.test_dir)))
        self.assertIsNone(sheet.cell(row=3, column=4).value) # No File 2 for unique

        # Validate "Summary" sheet
        self.assertIn("Summary", workbook.sheetnames)
        summary_sheet = workbook["Summary"]
        self.assertEqual(summary_sheet["A1"].value, "Metric")
        self.assertEqual(summary_sheet["B1"].value, "Value")
        self.assertEqual(summary_sheet["A2"].value, "Total Files to Keep")
        self.assertEqual(summary_sheet["B2"].value, 2) # 1 keeper from duplicate + 1 unique
        self.assertEqual(summary_sheet["A3"].value, "Total Size to Keep (Bytes)")
        self.assertEqual(summary_sheet["B3"].value, file1.stat().st_size + file3.stat().st_size)
        self.assertEqual(summary_sheet["A4"].value, "Total Size to Keep (GB)")
        # Check value type and approximate value
        self.assertIsInstance(float(summary_sheet["B4"].value), float)
        self.assertAlmostEqual(float(summary_sheet["B4"].value), (file1.stat().st_size + file3.stat().st_size) / (1024**3), places=5)

        # Clean up the created file
        output_file.unlink()


if __name__ == "__main__":
    unittest.main()