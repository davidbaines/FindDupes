import argparse
from pathlib import Path
import openpyxl
from tqdm import tqdm
from concurrent.futures import ThreadPoolExecutor, as_completed

from duplicate_file_finder import DEFAULT_REPORT_FILE


def get_files_from_report(report_path: Path, sheet_name: str) -> tuple[set[str], str]:
    """
    Reads an XLSX report and returns a tuple containing:
    - A set of all relative file paths found.
    - The actual name of the sheet that was read.
    """
    if not report_path.is_file():
        raise FileNotFoundError(f"Report file not found at '{report_path}'")

    print(f"Reading existing file paths from '{report_path.name}'...")
    workbook = openpyxl.load_workbook(report_path, read_only=True)

    actual_sheet_name = sheet_name
    if actual_sheet_name not in workbook.sheetnames:
        # Try to find a likely sheet name if the default isn't there
        if "Duplicates" in workbook.sheetnames:
            actual_sheet_name = "Duplicates"
            print(f"Info: Sheet '{sheet_name}' not found. Using existing sheet 'Duplicates'.")
        else:
            raise ValueError(f"Sheet '{sheet_name}' not found in the report.")

    sheet = workbook[actual_sheet_name]
    report_paths = set()
    # Iterate rows, skipping header
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # File paths start from the 3rd column (index 2)
        for cell_value in row[2:]:
            if cell_value:  # Check if cell is not empty
                report_paths.add(str(cell_value).strip().replace("\\", "/"))
    return report_paths, actual_sheet_name


def _get_disk_file_stat(path: Path, root_dir: Path) -> tuple[str, int] | None:
    """Helper for multithreading: gets stat for a single file, returns (relative_path, size) or None."""
    try:
        if path.is_file():
            stats = path.stat()
            if stats.st_size > 0:
                relative_path = str(path.relative_to(root_dir)).replace("\\", "/")
                return (relative_path, stats.st_size)
    except OSError:
        return None
    return None


def get_files_from_disk(root_dir: Path) -> dict[str, int]:
    """Scans the directory and returns a dict of {relative_path: size} for all non-empty files."""
    disk_files = {}
    print("Scanning disk for all current file paths...")
    paths_to_scan = list(root_dir.rglob("*"))
    with ThreadPoolExecutor() as executor:
        with tqdm(total=len(paths_to_scan), desc="Scanning disk") as pbar:
            futures = [executor.submit(_get_disk_file_stat, path, root_dir) for path in paths_to_scan]
            for future in as_completed(futures):
                result = future.result()
                if result:
                    path_str, size = result
                    disk_files[path_str] = size
                pbar.update(1)
    return disk_files


def append_to_report(report_path: Path, sheet_name: str, unique_files: list[tuple[str, int]]):
    """Appends a list of unique files (with sizes) to the specified sheet in the XLSX report."""
    print(f"Appending {len(unique_files)} new files to the report...")
    workbook = openpyxl.load_workbook(report_path)
    sheet = workbook[sheet_name]

    for file_path, size in tqdm(unique_files, desc="Updating report"):
        # Row format: [Keep, Size (Bytes), File 1]
        row_to_add = [1, size, file_path]
        sheet.append(row_to_add)

    workbook.save(report_path)
    print("Report successfully updated.")


def main():
    parser = argparse.ArgumentParser(
        description="Fast update of an existing XLSX report by appending unique files found on disk. "
        "This does NOT re-scan for duplicates and preserves existing report data."
    )
    parser.add_argument("folder", help="The root folder to scan for new unique files.")
    parser.add_argument(
        "-r",
        "--report-name",
        default=DEFAULT_REPORT_FILE,
        help=f"Name of the XLSX report file to update. Defaults to '{DEFAULT_REPORT_FILE}'.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Show which new files would be added without modifying the report.",
    )
    args = parser.parse_args()

    root_directory = Path(args.folder).resolve()
    report_path = root_directory / args.report_name
    sheet_name = "Files to Process"

    try:
        report_files, actual_sheet_name = get_files_from_report(report_path, sheet_name)
        disk_files_with_sizes = get_files_from_disk(root_directory)

        # Find paths on disk that are not in the report
        new_file_paths = sorted(list(disk_files_with_sizes.keys() - report_files))

        # Create the list of (path, size) tuples to add
        unique_files_to_add = [(path, disk_files_with_sizes[path]) for path in new_file_paths]

        print(f"\nFound {len(report_files)} files in the report.")
        print(f"Found {len(disk_files_with_sizes)} files on disk.")
        print(f"Found {len(unique_files_to_add)} new unique files to add.")

        if not unique_files_to_add:
            print("The report is already up-to-date. No new files found.")
        elif args.dry_run:
            print("\n--- DRY RUN: The following files would be added to the report ---")
            for path, size in unique_files_to_add:
                print(f"  - {path} ({size} bytes)")
        else:
            append_to_report(report_path, actual_sheet_name, unique_files_to_add)

    except (FileNotFoundError, ValueError) as e:
        print(f"\nERROR: {e}")


if __name__ == "__main__":
    main()
